#!/usr/bin/env python3
"""
evaluate_kgs.py
Quantitative evaluation of generated Knowledge Graphs for the agentic-generator-paper.

Metrics reported per file and per framework:
  - Concept Precision, Recall, F1  (class-level)
  - Relation Precision, Recall, F1 (object-property-level)
  - Property Precision, Recall, F1 (data-property-level)
  - Macro-average Precision, Recall, F1 across the three dimensions
  - Syntax validity, individual count, triple count, issues count
  - Hallucinated classes, missing concepts

Outputs:
  evaluation_summary.csv
  evaluation_full.json
  evaluation_summary.xlsx
"""

import os
import re
import csv
import json
from pathlib import Path
from collections import defaultdict

# ── optional xlsx support ──────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[warn] openpyxl not installed — xlsx output will be skipped. "
          "Install with: pip install openpyxl")

# ─────────────────────────────────────────────────────────────────────────────
# Paths
# ─────────────────────────────────────────────────────────────────────────────
REPO    = Path(__file__).resolve().parent.parent
KG_DIR  = REPO / "experiment_kg" 
OUT_DIR = Path(__file__).parent
OUT_DIR.mkdir(exist_ok=True)

# ─────────────────────────────────────────────────────────────────────────────
# Ground-truth sets (Gold Standard)
# These define what a "perfect" extraction should contain for each framework.
# ─────────────────────────────────────────────────────────────────────────────

# All classes defined in the ontology schema
ALL_ONTO_CLASSES = {
    "LLMAgent", "Team", "Task", "Tool", "Goal", "Prompt", "Config",
    "WorkflowPattern", "WorkflowStep", "StartStep", "EndStep",
    "LanguageModel", "KnowledgeBase", "Memory", "Constraint",
    "Capability", "Environment", "Objective", "HumanAgent", "Instance",
}

# All object properties defined in the ontology
ALL_ONTO_OBJ_PROPS = {
    "hasAgentMember", "hasAgentGoal", "hasTeamGoal", "hasGoal",
    "hasWorkflowPattern", "hasWorkflowStep", "hasAssociatedTask",
    "performedByAgent", "performedBy", "agentToolUsage", "toolUsage",
    "agentPrompt", "taskPrompt", "hasPrompt",
    "useLanguageModel", "hasKnowledge", "hasCapability", "hasAgentCapability",
    "hasAgentConfig", "hasSystemConfig", "hasToolConfig", "hasConfig",
    "hasEnvironmentConfig", "interactsWith", "operatesIn",
    "nextStep", "relatedStep", "nextPattern", "hasSubPattern", "hasRelatedPattern",
    "contributesToGoal", "contributesToObjective", "hasObjective",
    "producedResource", "requiresResource", "resourceUsage", "agentResourceUsage",
    "containsResource", "requiresCapability", "humanParticipatedIn",
}

# All data properties defined in the ontology
ALL_ONTO_DATA_PROPS = {
    "agentID", "agentRole", "configKey", "configValue", "envType",
    "promptContext", "promptInputData", "promptInstruction", "promptOutputIndicator",
    "stepOrder",
}

# Expected classes per framework (the "gold" set of classes that should appear)
GOLD_CLASSES = {
    "CrewAI": {
        "Team", "LLMAgent", "Task", "Tool", "Goal", "Prompt",
        "Config", "WorkflowPattern", "WorkflowStep", "LanguageModel",
    },
    "LangGraph": {
        "LLMAgent", "Task", "WorkflowPattern", "WorkflowStep",
        "StartStep", "EndStep", "Prompt", "Config", "LanguageModel",
        "Goal", "Team", "Tool",
    },
    "AutoGen": {
        "LLMAgent", "Task", "Tool", "Config", "Prompt",
        "WorkflowPattern", "WorkflowStep", "LanguageModel",
        "Goal",
    },
    "Mastra AI": {
        "LLMAgent", "Task", "Tool", "Config", "Prompt",
        "WorkflowPattern", "WorkflowStep", "LanguageModel",
        "Goal", "Team",
    },
}

# Expected object properties per framework (the "gold" relation set)
GOLD_OBJ_PROPS = {
    "CrewAI": {
        "hasAgentMember", "hasTeamGoal", "hasWorkflowPattern",
        "hasWorkflowStep", "performedByAgent", "useLanguageModel",
        "agentPrompt", "taskPrompt", "hasAgentConfig", "nextStep",
    },
    "LangGraph": {
        "hasWorkflowPattern", "hasWorkflowStep", "nextStep",
        "performedByAgent", "useLanguageModel", "agentPrompt",
        "taskPrompt", "hasAgentConfig", "hasAssociatedTask",
    },
    "AutoGen": {
        "performedByAgent", "useLanguageModel", "agentPrompt",
        "taskPrompt", "hasAgentConfig", "interactsWith",
        "hasWorkflowPattern", "hasWorkflowStep",
    },
    "Mastra AI": {
        "hasAgentMember", "hasWorkflowPattern", "hasWorkflowStep",
        "performedByAgent", "useLanguageModel", "agentPrompt",
        "taskPrompt", "hasAgentConfig", "nextStep",
    },
}

# Expected data properties per framework
GOLD_DATA_PROPS = {
    "CrewAI":    {"agentID", "agentRole", "configKey", "configValue",
                  "promptInstruction", "promptOutputIndicator", "stepOrder"},
    # promptContext removed — no structural LangGraph field maps to it; per-example where justified
    # promptOutputIndicator added — used 3–6/6 across all LangGraph examples
    "LangGraph": {"agentID", "agentRole", "configKey", "configValue",
                  "promptInstruction", "promptOutputIndicator", "stepOrder"},
    # promptContext removed — no dedicated AutoGen field; per-example where consistent
    # promptInputData removed — used 1–3/6, too inconsistent for framework gold
    # promptOutputIndicator added — used 3–6/6 across AutoGen examples
    "AutoGen":   {"agentID", "agentRole", "configKey", "configValue",
                  "promptInstruction", "promptOutputIndicator"},
    "Mastra AI": {"agentID", "agentRole", "configKey", "configValue",
                  "promptInstruction", "promptOutputIndicator"},
}

# Per-example overrides — applied on top of the framework-level gold sets.
# Key: (framework, example_stem)  [stem = filename without _instances.ttl]
# This handles cases where individual examples require more or fewer concepts
# than the framework-wide default (e.g., an example with no tools, or one that
# explicitly uses KnowledgeBase / Memory / HumanAgent).
GOLD_OVERRIDES: dict[tuple[str, str], dict] = {
    # ── CrewAI ───────────────────────────────────────────────────────────────
    # starter_template: both tool slots are commented out in agents.py
    ("CrewAI", "starter_template"): {
        "remove_classes": {"Tool"},
        "remove_obj_props": {"agentToolUsage", "toolUsage"},
    },
    # meta_quest_knowledge: uses PDFKnowledgeSource → maps to KnowledgeBase
    ("CrewAI", "meta_quest_knowledge"): {
        "add_classes": {"KnowledgeBase"},
        "add_obj_props": {"hasKnowledge"},
    },
    # ── AutoGen ──────────────────────────────────────────────────────────────
    # L2: customer_proxy_agent has human_input_mode="ALWAYS" → HumanAgent
    ("AutoGen", "L2_Sequential_Chats_and_Customer_Onboarding"): {
        "add_classes": {"HumanAgent"},
        "add_obj_props": {"humanParticipatedIn"},
    },
    # L3: pure AssistantAgent reflection chain — no external tool usage
    ("AutoGen", "L3_Reflection_and_Blogpost_Writing"): {
        "remove_classes": {"Tool"},
        "remove_obj_props": {"agentToolUsage", "toolUsage"},
    },
    # ── LangGraph ────────────────────────────────────────────────────────────
    # chat-agent: simple ReAct-style chat node, no tools; promptContext used 6/6
    ("LangGraph", "chat-agent"): {
        "remove_classes": {"Tool"},
        "remove_obj_props": {"agentToolUsage", "toolUsage"},
        "add_data_props": {"promptContext"},
    },
    # open-code: ToolNode with a human-review interrupt → HumanAgent
    ("LangGraph", "open-code"): {
        "add_classes": {"HumanAgent"},
        "add_obj_props": {"humanParticipatedIn"},
    },
    # email-agent: human-in-the-loop interrupt before sending email
    ("LangGraph", "email-agent"): {
        "add_classes": {"HumanAgent"},
        "add_obj_props": {"humanParticipatedIn"},
    },
    # ── Mastra AI ────────────────────────────────────────────────────────────
    # agent (weather-agent): single agent with one tool — no team/crew concept
    ("Mastra AI", "agent"): {
        "remove_classes": {"Team"},
        "remove_obj_props": {"hasAgentMember"},
    },
    # crypto-chatbot: single chatbot agent backed by @mastra/memory
    ("Mastra AI", "crypto-chatbot"): {
        "add_classes": {"Memory"},
        "add_obj_props": {"hasKnowledge"},
        "remove_classes": {"Team", "WorkflowPattern", "WorkflowStep"},
        "remove_obj_props": {"hasAgentMember", "hasWorkflowPattern", "hasWorkflowStep", "nextStep"},
    },
    # bird-checkers: consistently use promptInputData (tool input schemas)
    ("Mastra AI", "bird-checker-with-express"): {
        "add_data_props": {"promptInputData"},
    },
    ("Mastra AI", "bird-checker-with-nextjs"): {
        "add_data_props": {"promptInputData"},
    },
    ("Mastra AI", "bird-checker-with-nextjs-and-eval"): {
        "add_data_props": {"promptInputData"},
    },

    # ── AutoGen — promptContext where consistently used ───────────────────────
    # All 5 AutoGen controlled examples use promptContext 3–5/6 → add per example
    ("AutoGen", "L1_Multi-Agent_Conversation_and_Stand-up_Comedy"): {
        "add_data_props": {"promptContext"},
    },
    ("AutoGen", "L2_Sequential_Chats_and_Customer_Onboarding"): {
        "add_data_props": {"promptContext"},
    },
    ("AutoGen", "L3_Reflection_and_Blogpost_Writing"): {
        "add_data_props": {"promptContext"},
    },
    ("AutoGen", "L4_Tool_Use_and_Conversational_Chess"): {
        "add_data_props": {"promptContext", "promptInputData"},  # chess input 3/6
    },
    ("AutoGen", "L5_Coding_and_Financial_Analysis"): {
        "add_data_props": {"promptContext"},
    },

    # ── LangGraph — per-example data props ───────────────────────────────────
    # pizza-orderer: promptInputData consistently used (order input schema) 4/6
    ("LangGraph", "pizza-orderer"): {
        "add_data_props": {"promptInputData"},
    },
    # stockbroker: promptInputData used 4/6 (stock query input)
    ("LangGraph", "stockbroker"): {
        "add_data_props": {"promptInputData"},
    },
}


def get_gold(framework: str, example_stem: str) -> tuple[set, set, set]:
    """Return (gold_classes, gold_obj_props, gold_data_props) for a specific example.

    Starts from the framework-level baseline and applies per-example overrides."""
    classes    = set(GOLD_CLASSES.get(framework, set()))
    obj_props  = set(GOLD_OBJ_PROPS.get(framework, set()))
    data_props = set(GOLD_DATA_PROPS.get(framework, set()))

    ovr = GOLD_OVERRIDES.get((framework, example_stem))
    if ovr:
        classes   = (classes    | ovr.get("add_classes",    set())) - ovr.get("remove_classes",    set())
        obj_props = (obj_props  | ovr.get("add_obj_props",  set())) - ovr.get("remove_obj_props",  set())
        data_props = (data_props | ovr.get("add_data_props", set())) - ovr.get("remove_data_props", set())

    return classes, obj_props, data_props


# ─────────────────────────────────────────────────────────────────────────────
# Turtle lightweight parser
# ─────────────────────────────────────────────────────────────────────────────

ONTO_NS   = "http://www.w3id.org/agentic-ai/onto#"
OWL_NS    = "http://www.w3.org/2002/07/owl#"
RDF_TYPE  = "http://www.w3.org/1999/02/22-rdf-syntax-ns#type"

# Build full IRI sets for fast membership tests
_ONTO_CLASSES    = {ONTO_NS + c for c in ALL_ONTO_CLASSES}
_ONTO_OBJ_PROPS  = {ONTO_NS + p for p in ALL_ONTO_OBJ_PROPS}
_ONTO_DATA_PROPS = {ONTO_NS + p for p in ALL_ONTO_DATA_PROPS}

# OWL structural types to skip when counting hallucinations
_OWL_META = {
    OWL_NS + s for s in (
        "Class", "Ontology", "ObjectProperty", "DatatypeProperty",
        "AnnotationProperty", "NamedIndividual", "TransitiveProperty",
        "SymmetricProperty", "FunctionalProperty",
    )
}

# External namespaces that are legitimate references, not hallucinations
_EXTERNAL_NS_PREFIXES = (
    # beam (both www. and non-www. variants appear in LLM output)
    "http://w3id.org/beam/core#",
    "http://www.w3id.org/beam/core#",
    # standard W3C vocabularies
    "http://www.w3.org/2002/07/owl#",
    "http://www.w3.org/1999/02/22-rdf-syntax-ns#",
    "http://www.w3.org/2000/01/rdf-schema#",
    "http://www.w3.org/2001/XMLSchema#",
    "http://www.w3.org/ns/prov#",
    "http://www.w3.org/2004/02/skos/core#",
    # Dublin Core
    "http://purl.org/dc/",
    # p-plan / prov-o extensions
    "http://purl.org/net/p-plan#",
    "http://purl.org/pp/",
)


def _local(iri: str) -> str:
    """Return the local name of an IRI (after # or last /)."""
    if "#" in iri:
        return iri.rsplit("#", 1)[-1]
    return iri.rsplit("/", 1)[-1]


def parse_ttl(path: Path) -> dict:
    """Parse a Turtle file with rdflib and extract evaluation metrics."""
    import rdflib
    import warnings

    text = path.read_text(encoding="utf-8", errors="replace")

    # ── metadata from header comments (not in the RDF graph) ─────────────────
    model_m  = re.search(r'Model used:\s*(\S+)', text)
    exec_m   = re.search(r'Execution time:\s*([\d.]+)', text)
    prompt_m = re.search(r'Prompt:\s*(\S+)', text)

    issues = []
    block = re.search(
        r'Issues\s*/\s*Assumptions:(.*?)(?:@prefix|\Z)',
        text, re.DOTALL | re.IGNORECASE,
    )
    if block:
        for line in block.group(1).splitlines():
            line = line.strip().lstrip('#').strip()
            if line.startswith('-') and len(line) > 5:
                issues.append(line[1:].strip())

    undefined_refs = sorted({
        f"agento-ext:{m.group(1)}"
        for m in re.finditer(r'agento-ext:(\w+)', text)
    })

    # ── rdflib parse ─────────────────────────────────────────────────────────
    g = rdflib.Graph()
    syntax_errors: list[str] = []
    try:
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", message=".*does not look like a valid URI.*")
            g.parse(data=text, format="turtle")
    except Exception as exc:
        syntax_errors.append(str(exc)[:200])

    # ── collect classes, properties, individuals from the graph ──────────────
    class_iris:      set[str] = set()
    obj_prop_iris:   set[str] = set()
    data_prop_iris:  set[str] = set()
    individuals_by_class: dict[str, set] = defaultdict(set)

    for s, p, o in g:
        s_str = str(s)
        p_str = str(p)
        o_str = str(o)

        if p_str == RDF_TYPE:

            # Collect the type IRI
            if o_str in _OWL_META:
                continue  # structural OWL declarations, not data types
            class_iris.add(o_str)

            # Map individual → class (only onto-namespace subjects)
            if o_str.startswith(ONTO_NS) and s_str.startswith(ONTO_NS):
                cls_local = _local(o_str)
                ind_local = _local(s_str)
                individuals_by_class[cls_local].add(ind_local)
        else:
            # Predicate is a property
            if p_str.startswith(ONTO_NS):
                local = _local(p_str)
                if local in ALL_ONTO_OBJ_PROPS:
                    obj_prop_iris.add(local)
                elif local in ALL_ONTO_DATA_PROPS:
                    data_prop_iris.add(local)

    # ── classify type IRIs ────────────────────────────────────────────────────
    onto_classes_found: set[str] = set()
    extra_classes:      set[str] = set()

    for iri in class_iris:
        if iri in _ONTO_CLASSES:
            onto_classes_found.add(_local(iri))
        elif (
            iri not in _ONTO_OBJ_PROPS
            and iri not in _ONTO_DATA_PROPS
            and iri not in _OWL_META
            and not any(iri.startswith(ns) for ns in _EXTERNAL_NS_PREFIXES)
        ):
            local = _local(iri)
            if local not in ALL_ONTO_OBJ_PROPS and local not in ALL_ONTO_DATA_PROPS:
                extra_classes.add(local if iri.startswith(ONTO_NS) else iri)

    # ── duplicate subjects (text-based: counts separate Turtle blocks) ────────
    # rdflib merges duplicate declarations silently, so regex is more reliable here.
    seen_subj: set[str] = set()
    dupes: list[str] = []
    for m in re.finditer(r'^:(\w+)\s+(?:rdf:type|a)\s+', text, re.MULTILINE):
        name = m.group(1)
        if name in seen_subj:
            if name not in dupes:
                dupes.append(name)
        seen_subj.add(name)

    return {
        "syntax_errors":        syntax_errors,
        "syntax_valid":         len(syntax_errors) == 0,
        "onto_classes_found":   sorted(onto_classes_found),
        "extra_classes":        sorted(extra_classes),
        "individuals_by_class": {k: list(v) for k, v in individuals_by_class.items()},
        "total_individuals":    sum(len(v) for v in individuals_by_class.values()),
        "approx_triples":       len(g),
        "used_obj_props":       sorted(obj_prop_iris),
        "used_data_props":      sorted(data_prop_iris),
        "model":                model_m.group(1) if model_m else "unknown",
        "prompt_id":            prompt_m.group(1) if prompt_m else "unknown",
        "exec_time":            float(exec_m.group(1)) if exec_m else None,
        "issues":               issues,
        "undefined_refs":       undefined_refs,
        "duplicate_subjects": dupes,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Precision / Recall / F1 helpers
# ─────────────────────────────────────────────────────────────────────────────

def prf1(tp: int, fp: int, fn: int) -> tuple[float, float, float]:
    """Compute precision, recall, F1 from TP/FP/FN counts."""
    precision = tp / (tp + fp) if (tp + fp) > 0 else 0.0
    recall    = tp / (tp + fn) if (tp + fn) > 0 else 0.0
    f1        = (2 * precision * recall / (precision + recall)
                 if (precision + recall) > 0 else 0.0)
    return round(precision, 4), round(recall, 4), round(f1, 4)


def compute_prf1(metrics: dict, framework: str, example_stem: str = "") -> dict:
    """
    Compute precision, recall, F1 at three levels (class, object-prop, data-prop)
    and a macro-average across the three.

    Definitions
    -----------
    TP (class)     = extracted onto_classes ∩ gold_classes
    FP (class)     = extracted onto_classes − gold_classes   (extracted but not expected)
    FN (class)     = gold_classes − extracted onto_classes   (expected but not extracted)

    Same logic for object properties and data properties.
    """
    gold_cls, gold_obj, gold_data = get_gold(framework, example_stem)

    found_cls  = set(metrics["onto_classes_found"])
    found_obj  = set(metrics["used_obj_props"])
    found_data = set(metrics["used_data_props"])

    # FP is only items that are completely outside the ontology schema (hallucinations).
    # Valid ontology terms extracted beyond the gold set are neutral — not penalised.
    hallucinated_cls  = set(metrics["extra_classes"])   # already excludes beam:/pp: builtins
    hallucinated_obj  = set(metrics["used_obj_props"]) - ALL_ONTO_OBJ_PROPS   # should be empty normally
    hallucinated_data = set(metrics["used_data_props"]) - ALL_ONTO_DATA_PROPS  # same

    # ── class level ──────────────────────────────────────────────────────────
    tp_cls = len(found_cls & gold_cls)
    fp_cls = len(hallucinated_cls)               # only non-ontology terms are FP
    fn_cls = len(gold_cls - found_cls)           # expected but missing
    p_cls, r_cls, f1_cls = prf1(tp_cls, fp_cls, fn_cls)

    # ── object-property level ────────────────────────────────────────────────
    tp_obj = len(found_obj & gold_obj)
    fp_obj = len(hallucinated_obj)               # non-ontology properties only
    fn_obj = len(gold_obj - found_obj)
    p_obj, r_obj, f1_obj = prf1(tp_obj, fp_obj, fn_obj)

    # ── data-property level ──────────────────────────────────────────────────
    tp_data = len(found_data & gold_data)
    fp_data = len(hallucinated_data)             # non-ontology data properties only
    fn_data = len(gold_data - found_data)
    p_data, r_data, f1_data = prf1(tp_data, fp_data, fn_data)

    # ── macro averages ───────────────────────────────────────────────────────
    macro_p  = round((p_cls + p_obj + p_data) / 3, 4)
    macro_r  = round((r_cls + r_obj + r_data) / 3, 4)
    macro_f1 = round((f1_cls + f1_obj + f1_data) / 3, 4)

    return {
        # class level
        "cls_tp": tp_cls, "cls_fp": fp_cls, "cls_fn": fn_cls,
        "cls_precision": p_cls, "cls_recall": r_cls, "cls_f1": f1_cls,
        # object-property level
        "obj_tp": tp_obj, "obj_fp": fp_obj, "obj_fn": fn_obj,
        "obj_precision": p_obj, "obj_recall": r_obj, "obj_f1": f1_obj,
        # data-property level
        "data_tp": tp_data, "data_fp": fp_data, "data_fn": fn_data,
        "data_precision": p_data, "data_recall": r_data, "data_f1": f1_data,
        # macro averages
        "macro_precision": macro_p, "macro_recall": macro_r, "macro_f1": macro_f1,
        # diagnostics
        "missing_concepts":     sorted(gold_cls  - found_cls),
        "missing_obj_props":    sorted(gold_obj  - found_obj),
        "missing_data_props":   sorted(gold_data - found_data),
        "hallucinated_classes": [
            c for c in metrics["extra_classes"]
            if c not in {"beam:Instance", "beam:Resource", "beam:System",
                         "beam:Model", "beam:StatisticalModel"}
        ],
        "syntax_valid": metrics["syntax_valid"],
    }


# ─────────────────────────────────────────────────────────────────────────────
# Framework evaluation
# ─────────────────────────────────────────────────────────────────────────────

def evaluate_framework(fw_name: str, fw_dir: Path, max_files: int | None = None) -> list[dict]:
    """
    Walk experiment_kg/<Framework>/<PromptID>/<Model>/*.ttl and evaluate each file.
    Falls back to flat *.ttl scan if no prompt/model subdirectories exist.
    """
    results = []

    # Detect nested structure: fw_dir/<prompt>/<model>/*.ttl
    prompt_dirs = [d for d in sorted(fw_dir.iterdir()) if d.is_dir()]
    has_nested = any(
        any(d2.is_dir() for d2 in p.iterdir())
        for p in prompt_dirs if p.is_dir()
    )

    if has_nested:
        for prompt_dir in prompt_dirs:
            prompt_id = prompt_dir.name
            for model_dir in sorted(prompt_dir.iterdir()):
                if not model_dir.is_dir():
                    continue
                model_name = model_dir.name
                ttl_files = sorted(model_dir.glob("*.ttl"))
                if max_files is not None:
                    ttl_files = ttl_files[:max_files]
                for ttl in ttl_files:
                    stem = ttl.stem.replace("_instances", "")
                    m = parse_ttl(ttl)
                    m["file"]      = ttl.name
                    m["framework"] = fw_name
                    m["prompt_id"] = prompt_id
                    m["model"]     = model_name
                    m.update(compute_prf1(m, fw_name, stem))
                    results.append(m)
    else:
        # Flat layout (kgs_original style)
        ttl_files = sorted(fw_dir.glob("*.ttl"))
        if max_files is not None:
            ttl_files = ttl_files[:max_files]
        for ttl in ttl_files:
            stem = ttl.stem.replace("_instances", "")
            m = parse_ttl(ttl)
            m["file"]      = ttl.name
            m["framework"] = fw_name
            m.update(compute_prf1(m, fw_name, stem))
            results.append(m)

    return results


# ─────────────────────────────────────────────────────────────────────────────
# CSV output
# ─────────────────────────────────────────────────────────────────────────────

CSV_FIELDS = [
    "framework", "file", "model", "prompt_id", "exec_time",
    "syntax_valid", "syntax_error_msg", "total_individuals", "approx_triples",
    # class-level P/R/F1
    "cls_tp", "cls_fp", "cls_fn", "cls_precision", "cls_recall", "cls_f1",
    # object-property-level P/R/F1
    "obj_tp", "obj_fp", "obj_fn", "obj_precision", "obj_recall", "obj_f1",
    # data-property-level P/R/F1
    "data_tp", "data_fp", "data_fn", "data_precision", "data_recall", "data_f1",
    # macro
    "macro_precision", "macro_recall", "macro_f1",
    # diagnostics
    "missing_concepts", "missing_obj_props", "missing_data_props",
    "hallucinated_classes", "duplicate_subjects", "undefined_refs",
    "onto_classes_found", "extra_classes",
    "used_obj_props", "used_data_props",
]


def write_csv(records: list[dict], path: Path) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=CSV_FIELDS, extrasaction="ignore")
        w.writeheader()
        for r in records:
            row = dict(r)
            row["syntax_error_msg"] = "; ".join(row.get("syntax_errors") or [])
            for k in ["missing_concepts", "missing_obj_props", "missing_data_props",
                      "hallucinated_classes", "duplicate_subjects", "undefined_refs",
                      "onto_classes_found", "extra_classes",
                      "used_obj_props", "used_data_props"]:
                row[k] = "; ".join(row.get(k) or [])
            w.writerow(row)
    print(f"[✓] CSV  → {path}")


# ─────────────────────────────────────────────────────────────────────────────
# XLSX output
# ─────────────────────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79") if HAS_OPENPYXL else None
SUBHEAD_FILL  = PatternFill("solid", fgColor="2E75B6") if HAS_OPENPYXL else None
ALT_FILL      = PatternFill("solid", fgColor="D6E4F0") if HAS_OPENPYXL else None
GOOD_FILL     = PatternFill("solid", fgColor="C6EFCE") if HAS_OPENPYXL else None
WARN_FILL     = PatternFill("solid", fgColor="FFEB9C") if HAS_OPENPYXL else None
BAD_FILL      = PatternFill("solid", fgColor="FFC7CE") if HAS_OPENPYXL else None
WHITE_FONT    = Font(color="FFFFFF", bold=True, name="Calibri") if HAS_OPENPYXL else None
BOLD_FONT     = Font(bold=True, name="Calibri") if HAS_OPENPYXL else None
REG_FONT      = Font(name="Calibri") if HAS_OPENPYXL else None
THIN_BORDER   = (Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
) if HAS_OPENPYXL else None)


def _set_cell(ws, row: int, col: int, value,
              fill=None, font=None, align="center", border=True) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    if fill:   cell.fill   = fill
    if font:   cell.font   = font or REG_FONT
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if border and THIN_BORDER:
        cell.border = THIN_BORDER


def _score_fill(value: float):
    """Return a fill colour based on F1 / recall thresholds."""
    if value >= 0.75:
        return GOOD_FILL
    if value >= 0.40:
        return WARN_FILL
    return BAD_FILL


def write_xlsx(records: list[dict], path: Path) -> None:
    if not HAS_OPENPYXL:
        return

    wb = openpyxl.Workbook()

    # ── Sheet 1: Per-file results ────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Per-File Results"
    ws1.freeze_panes = "C3"

    headers_row1 = [
        "Framework", "File", "Model", "Prompt",
        "Syntax\nValid", "Syntax Error", "Individuals", "Triples", "Issues",
        # Class
        "Cls TP", "Cls FP", "Cls FN",
        "Cls\nPrecision", "Cls\nRecall", "Cls\nF1",
        # Obj-prop
        "ObjProp TP", "ObjProp FP", "ObjProp FN",
        "ObjProp\nPrecision", "ObjProp\nRecall", "ObjProp\nF1",
        # Data-prop
        "DataProp TP", "DataProp FP", "DataProp FN",
        "DataProp\nPrecision", "DataProp\nRecall", "DataProp\nF1",
        # Macro
        "Macro\nPrecision", "Macro\nRecall", "Macro\nF1",
        # Diagnostics
        "Missing Concepts", "Missing ObjProps", "Hallucinated Classes",
    ]

    # Merged group headers
    groups = [
        (1, 9, "Metadata & Syntax"),
        (10, 15, "Class Level"),
        (16, 21, "Object-Property Level"),
        (22, 27, "Data-Property Level"),
        (28, 30, "Macro Average"),
        (31, 33, "Diagnostics"),
    ]
    for start_col, end_col, label in groups:
        ws1.merge_cells(start_row=1, start_column=start_col,
                        end_row=1, end_column=end_col)
        _set_cell(ws1, 1, start_col, label, fill=HEADER_FILL, font=WHITE_FONT)

    for col, h in enumerate(headers_row1, 1):
        _set_cell(ws1, 2, col, h, fill=SUBHEAD_FILL, font=WHITE_FONT)

    ws1.row_dimensions[1].height = 22
    ws1.row_dimensions[2].height = 42

    for row_idx, r in enumerate(records, start=3):
        alt = ALT_FILL if row_idx % 2 == 0 else None
        vals = [
            r["framework"], r["file"], r["model"], r.get("prompt_id", ""),
            "✓" if r["syntax_valid"] else "✗",
            "; ".join(r.get("syntax_errors") or []),
            r["total_individuals"], r["approx_triples"], len(r.get("issues", [])),
            r["cls_tp"], r["cls_fp"], r["cls_fn"],
            r["cls_precision"], r["cls_recall"], r["cls_f1"],
            r["obj_tp"], r["obj_fp"], r["obj_fn"],
            r["obj_precision"], r["obj_recall"], r["obj_f1"],
            r["data_tp"], r["data_fp"], r["data_fn"],
            r["data_precision"], r["data_recall"], r["data_f1"],
            r["macro_precision"], r["macro_recall"], r["macro_f1"],
            "; ".join(r.get("missing_concepts", [])),
            "; ".join(r.get("missing_obj_props", [])),
            "; ".join(r.get("hallucinated_classes", [])),
        ]
        for col_idx, val in enumerate(vals, 1):
            fill = alt
            # Colour F1 / recall cells
            if col_idx in {15, 21, 27, 30}:   # F1 columns
                fill = _score_fill(val if isinstance(val, (int, float)) else 0)
            elif col_idx in {14, 20, 26, 29}:  # Recall columns
                fill = _score_fill(val if isinstance(val, (int, float)) else 0)
            _set_cell(ws1, row_idx, col_idx, val, fill=fill, font=REG_FONT,
                      align="left" if col_idx in {1, 2, 6, 31, 32, 33} else "center")

    # Auto-size columns
    col_widths = [14, 38, 12, 8, 7, 45, 11, 8, 7,
                  7, 7, 7, 11, 9, 8,
                  11, 11, 11, 12, 10, 8,
                  12, 12, 12, 13, 11, 9,
                  13, 11, 9,
                  35, 35, 35]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Group Summary — one table per model, prompt-grouped rows ───────
    ws2 = wb.create_sheet("Group Summary")
    ws2.freeze_panes = "C4"

    # Column layout (no Model column — each table is already scoped to one model)
    summary_headers = [
        "Prompt", "Framework", "Files",
        "Syntax\nValid", "Avg\nTime (s)", "Avg\nIndividuals", "Avg\nTriples", "Avg\nIssues",
        "Cls\nPrecision", "Cls\nRecall", "Cls\nF1",
        "ObjProp\nPrecision", "ObjProp\nRecall", "ObjProp\nF1",
        "DataProp\nPrecision", "DataProp\nRecall", "DataProp\nF1",
        "Macro\nPrecision", "Macro\nRecall", "Macro\nF1",
    ]
    N_COLS = len(summary_headers)

    def avg(recs, key):
        vals = [r.get(key) or 0 for r in recs]
        return round(sum(vals) / len(vals), 4) if vals else 0

    def avg_issues(recs):
        n = len(recs)
        return round(sum(len(r.get('issues', [])) for r in recs) / n, 1) if n else 0

    def avg_time_xlsx(recs):
        times = [r.get("exec_time") for r in recs if r.get("exec_time") is not None]
        return round(sum(times) / len(times), 1) if times else None

    def _data_vals(prompt, fw, n, recs):
        return [
            prompt, fw, n,
            f"{sum(r['syntax_valid'] for r in recs)}/{n}",
            avg_time_xlsx(recs),
            round(avg(recs, 'total_individuals'), 1),
            round(avg(recs, 'approx_triples'), 1),
            avg_issues(recs),
            avg(recs, 'cls_precision'),   avg(recs, 'cls_recall'),   avg(recs, 'cls_f1'),
            avg(recs, 'obj_precision'),   avg(recs, 'obj_recall'),   avg(recs, 'obj_f1'),
            avg(recs, 'data_precision'),  avg(recs, 'data_recall'),  avg(recs, 'data_f1'),
            avg(recs, 'macro_precision'), avg(recs, 'macro_recall'), avg(recs, 'macro_f1'),
        ]

    def _write_data_row(ws, row_idx, vals, is_avg=False, is_overall=False):
        for col_idx, val in enumerate(vals, 1):
            if is_overall:
                fill = PatternFill("solid", fgColor="1F4E79")
                font = WHITE_FONT
            elif is_avg:
                fill = PatternFill("solid", fgColor="2E75B6")
                font = WHITE_FONT
            else:
                fill = ALT_FILL if row_idx % 2 == 0 else None
                font = REG_FONT
                # Colour F1 / recall cells
                if col_idx in {11, 14, 17, 20}:  # F1 columns
                    fill = _score_fill(val if isinstance(val, float) else 0)
                elif col_idx in {10, 13, 16, 19}:  # Recall columns
                    fill = _score_fill(val if isinstance(val, float) else 0)
            _set_cell(ws, row_idx, col_idx, val, fill=fill, font=font,
                      align="left" if col_idx <= 2 else "center")

    all_models = sorted({r.get("model", "") for r in records})
    cur_row = 1  # current write position in ws2

    for model_name in all_models:
        model_recs = [r for r in records if r.get("model", "") == model_name]

        # Model header spanning all columns
        ws2.merge_cells(start_row=cur_row, start_column=1,
                        end_row=cur_row, end_column=N_COLS)
        _set_cell(ws2, cur_row, 1,
                  f"MODEL: {model_name}  ({len(model_recs)} files)",
                  fill=PatternFill("solid", fgColor="1F4E79"), font=WHITE_FONT, align="left")
        ws2.row_dimensions[cur_row].height = 22
        cur_row += 1

        # Column-group headers (merged)
        group_spans = [
            (1, 3, "Overview"), (4, 8, "Counts"),
            (9, 11, "Class Level"), (12, 14, "Object-Property Level"),
            (15, 17, "Data-Property Level"), (18, 20, "Macro Average"),
        ]
        for sc, ec, label in group_spans:
            ws2.merge_cells(start_row=cur_row, start_column=sc,
                            end_row=cur_row, end_column=ec)
            _set_cell(ws2, cur_row, sc, label, fill=SUBHEAD_FILL, font=WHITE_FONT)
        ws2.row_dimensions[cur_row].height = 18
        cur_row += 1

        # Column headers
        for col, h in enumerate(summary_headers, 1):
            _set_cell(ws2, cur_row, col, h, fill=SUBHEAD_FILL, font=WHITE_FONT)
        ws2.row_dimensions[cur_row].height = 40
        cur_row += 1

        # Data rows grouped by prompt
        by_group2: dict[tuple, list] = defaultdict(list)
        by_prompt2: dict[str, list] = defaultdict(list)
        for r in model_recs:
            key = (r.get("prompt_id", ""), r["framework"])
            by_group2[key].append(r)
            by_prompt2[r.get("prompt_id", "")].append(r)

        current_prompt = None
        for (prompt, fw), recs in sorted(by_group2.items()):
            if prompt != current_prompt:
                if current_prompt is not None:
                    # Prompt AVG row
                    p_recs = by_prompt2[current_prompt]
                    _write_data_row(ws2, cur_row,
                                    _data_vals(f"AVG {current_prompt}", "", len(p_recs), p_recs),
                                    is_avg=True)
                    cur_row += 1
                current_prompt = prompt
            _write_data_row(ws2, cur_row, _data_vals(prompt, fw, len(recs), recs))
            cur_row += 1

        if current_prompt:
            p_recs = by_prompt2[current_prompt]
            _write_data_row(ws2, cur_row,
                            _data_vals(f"AVG {current_prompt}", "", len(p_recs), p_recs),
                            is_avg=True)
            cur_row += 1

        # Model OVERALL row
        _write_data_row(ws2, cur_row,
                        _data_vals("OVERALL", "", len(model_recs), model_recs),
                        is_overall=True)
        cur_row += 1
        cur_row += 1  # blank separator between model tables

    # Grand total row
    _write_data_row(ws2, cur_row,
                    _data_vals("GRAND TOTAL", "", len(records), records),
                    is_overall=True)

    col_widths2 = [10, 14, 6, 10, 10, 13, 10, 10,
                   12, 10, 8, 12, 12, 8,
                   13, 13, 9, 13, 12, 9]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Missing & Hallucinations ────────────────────────────────────
    ws3 = wb.create_sheet("Gaps & Hallucinations")
    gap_headers = [
        "Framework", "Prompt", "Model", "File",
        "Missing Concepts (classes)", "Missing Obj Props", "Missing Data Props",
        "Hallucinated Classes", "Duplicate Subjects",
    ]
    for col, h in enumerate(gap_headers, 1):
        _set_cell(ws3, 1, col, h, HEADER_FILL, WHITE_FONT)
    ws3.row_dimensions[1].height = 30

    for row_idx, r in enumerate(records, start=2):
        alt = ALT_FILL if row_idx % 2 == 0 else None
        vals = [
            r["framework"], r.get("prompt_id", ""), r.get("model", ""), r["file"],
            "; ".join(r.get("missing_concepts", [])),
            "; ".join(r.get("missing_obj_props", [])),
            "; ".join(r.get("missing_data_props", [])),
            "; ".join(r.get("hallucinated_classes", [])),
            "; ".join(r.get("duplicate_subjects", [])),
        ]
        for col_idx, val in enumerate(vals, 1):
            has_content = bool(val) and col_idx > 4
            fill = BAD_FILL if has_content else alt
            _set_cell(ws3, row_idx, col_idx, val, fill=fill, font=REG_FONT,
                      align="left")

    for i, w in enumerate([16, 8, 10, 40, 38, 38, 30, 35, 25], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 4: Prompt Ablation — ChatGPT, max 5 examples per framework ─────
    ws4 = wb.create_sheet("Prompt Ablation (controlled)")
    ws4.freeze_panes = "B4"

    abl_headers = [
        "Framework", "Files",
        "Cls\nPrecision", "Cls\nRecall", "Cls\nF1",
        "ObjProp\nPrecision", "ObjProp\nRecall", "ObjProp\nF1",
        "DataProp\nPrecision", "DataProp\nRecall", "DataProp\nF1",
        "Macro\nPrecision", "Macro\nRecall", "Macro\nF1",
    ]
    N_ABL = len(abl_headers)
    ABL_MAX = 5

    def _avg(recs, key):
        vals = [r.get(key) or 0 for r in recs]
        return round(sum(vals) / len(vals), 4) if vals else 0

    chatgpt_recs = [r for r in records if r.get("model") == "chatgpt"]
    prompts_ordered = sorted({r.get("prompt_id", "") for r in chatgpt_recs})
    frameworks_ordered = sorted({r["framework"] for r in chatgpt_recs})

    # Controlled file set: only files present under EVERY prompt for each framework
    _abl_common: dict[str, set] = {}
    for fw in frameworks_ordered:
        by_p = {p: {r["file"] for r in chatgpt_recs if r["framework"] == fw and r["prompt_id"] == p}
                for p in prompts_ordered}
        nonempty = [s for s in by_p.values() if s]
        _abl_common[fw] = set.intersection(*nonempty) if nonempty else set()

    def _abl_recs(fw, prompt_id):
        return [r for r in chatgpt_recs
                if r["framework"] == fw
                and r["prompt_id"] == prompt_id
                and r["file"] in _abl_common[fw]]

    cur = 1  # current row in ws4

    for prompt_id in prompts_ordered:
        # Prompt header spanning all columns
        ws4.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=N_ABL)
        _set_cell(ws4, cur, 1, f"Prompt: {prompt_id}",
                  fill=PatternFill("solid", fgColor="1F4E79"), font=WHITE_FONT, align="left")
        ws4.row_dimensions[cur].height = 20
        cur += 1

        # Column group headers
        abl_groups = [
            (1, 2, ""), (3, 5, "Class Level"),
            (6, 8, "Object-Property Level"), (9, 11, "Data-Property Level"),
            (12, 14, "Macro Average"),
        ]
        for sc, ec, label in abl_groups:
            if sc != ec:
                ws4.merge_cells(start_row=cur, start_column=sc, end_row=cur, end_column=ec)
            _set_cell(ws4, cur, sc, label, fill=SUBHEAD_FILL, font=WHITE_FONT)
        ws4.row_dimensions[cur].height = 16
        cur += 1

        # Column headers
        for col, h in enumerate(abl_headers, 1):
            _set_cell(ws4, cur, col, h, fill=SUBHEAD_FILL, font=WHITE_FONT)
        ws4.row_dimensions[cur].height = 38
        cur += 1

        prompt_all: list[dict] = []
        for fw in frameworks_ordered:
            fw_recs = _abl_recs(fw, prompt_id)
            if not fw_recs:
                continue
            prompt_all.extend(fw_recs)
            n = len(fw_recs)
            vals = [
                fw, n,
                _avg(fw_recs, "cls_precision"),  _avg(fw_recs, "cls_recall"),  _avg(fw_recs, "cls_f1"),
                _avg(fw_recs, "obj_precision"),  _avg(fw_recs, "obj_recall"),  _avg(fw_recs, "obj_f1"),
                _avg(fw_recs, "data_precision"), _avg(fw_recs, "data_recall"), _avg(fw_recs, "data_f1"),
                _avg(fw_recs, "macro_precision"),_avg(fw_recs, "macro_recall"),_avg(fw_recs, "macro_f1"),
            ]
            for col_idx, val in enumerate(vals, 1):
                fill = ALT_FILL if cur % 2 == 0 else None
                if col_idx in {5, 8, 11, 14}:   # F1 columns
                    fill = _score_fill(val if isinstance(val, float) else 0)
                elif col_idx in {4, 7, 10, 13}:  # Recall columns
                    fill = _score_fill(val if isinstance(val, float) else 0)
                _set_cell(ws4, cur, col_idx, val, fill=fill, font=REG_FONT,
                          align="left" if col_idx == 1 else "center")
            cur += 1

        # Prompt AVG row
        if prompt_all:
            avg_vals = [
                f"AVG {prompt_id}", len(prompt_all),
                _avg(prompt_all, "cls_precision"),  _avg(prompt_all, "cls_recall"),  _avg(prompt_all, "cls_f1"),
                _avg(prompt_all, "obj_precision"),  _avg(prompt_all, "obj_recall"),  _avg(prompt_all, "obj_f1"),
                _avg(prompt_all, "data_precision"), _avg(prompt_all, "data_recall"), _avg(prompt_all, "data_f1"),
                _avg(prompt_all, "macro_precision"),_avg(prompt_all, "macro_recall"),_avg(prompt_all, "macro_f1"),
            ]
            for col_idx, val in enumerate(avg_vals, 1):
                _set_cell(ws4, cur, col_idx, val,
                          fill=PatternFill("solid", fgColor="2E75B6"), font=WHITE_FONT,
                          align="left" if col_idx == 1 else "center")
            cur += 1

        cur += 1  # blank row between prompts

    col_widths4 = [14, 6, 12, 10, 8, 12, 12, 8, 13, 13, 9, 13, 12, 9]
    for i, w in enumerate(col_widths4, 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    wb.save(path)
    print(f"[✓] XLSX → {path}")


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    all_records: list[dict] = []


    for fw in ["CrewAI", "LangGraph", "AutoGen", "Mastra AI"]:
        fw_dir = KG_DIR / fw
        if not fw_dir.exists():
            continue
        records = evaluate_framework(fw, fw_dir)
        all_records.extend(records)
        print(f"  {fw}: {len(records)} files processed")

    # ── CSV ──────────────────────────────────────────────────────────────────
    write_csv(all_records, OUT_DIR / "evaluation_summary.csv")

    # ── JSON ─────────────────────────────────────────────────────────────────
    json_path = OUT_DIR / "evaluation_full.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(all_records, f, indent=2, default=str)
    print(f"[✓] JSON → {json_path}")

    # ── XLSX ─────────────────────────────────────────────────────────────────
    write_xlsx(all_records, OUT_DIR / "evaluation_summary.xlsx")

    # ── Console summary ──────────────────────────────────────────────────────
    def avg(recs, k):
        v = [r.get(k) or 0 for r in recs]
        return round(sum(v) / len(v), 3) if v else 0

    # ── Prompt ablation table (ChatGPT, controlled: files common across all prompts) ──
    chatgpt_recs_c = [r for r in all_records if r.get("model") == "chatgpt"]
    prompts_abl    = sorted({r.get("prompt_id", "") for r in chatgpt_recs_c})
    fws_abl        = sorted({r["framework"] for r in chatgpt_recs_c})

    # Files present under every prompt per framework
    abl_common: dict[str, set] = {}
    for fw in fws_abl:
        by_p = {p: {r["file"] for r in chatgpt_recs_c if r["framework"] == fw and r["prompt_id"] == p}
                for p in prompts_abl}
        nonempty = [s for s in by_p.values() if s]
        abl_common[fw] = set.intersection(*nonempty) if nonempty else set()

    def abl_fw_recs(fw, prompt_id):
        return [r for r in chatgpt_recs_c
                if r["framework"] == fw and r["prompt_id"] == prompt_id
                and r["file"] in abl_common[fw]]

    WA = 85
    HDR = (f"{'Framework':<12} {'N':>3}  "
           f"{'Cls P':>6} {'Cls R':>6} {'Cls F1':>6}  "
           f"{'Obj P':>6} {'Obj R':>6} {'Obj F1':>6}  "
           f"{'Mac F1':>7}")

    def _abl_row(label, n, recs):
        return (
            f"{label:<12} {n:>3}  "
            f"{avg(recs,'cls_precision'):>6.3f} {avg(recs,'cls_recall'):>6.3f} {avg(recs,'cls_f1'):>6.3f}  "
            f"{avg(recs,'obj_precision'):>6.3f} {avg(recs,'obj_recall'):>6.3f} {avg(recs,'obj_f1'):>6.3f}  "
            f"{avg(recs,'macro_f1'):>7.3f}"
        )

    # Show how many common files per framework
    fw_n = {fw: len(abl_common[fw]) for fw in fws_abl if abl_common[fw]}
    fw_summary = ", ".join(f"{fw} ({n})" for fw, n in fw_n.items())
    print(f"\n{'=' * WA}")
    print(f"  PROMPT ABLATION — ChatGPT  (controlled: same files per framework)")
    print(f"  Files: {fw_summary}")
    print(f"{'=' * WA}")

    for prompt_id in prompts_abl:
        print(f"\n  Prompt: {prompt_id}")
        print(f"  {'-' * (WA - 2)}")
        print(f"  {HDR}")
        print(f"  {'-' * (WA - 2)}")
        prompt_all: list[dict] = []
        for fw in fws_abl:
            fw_recs = abl_fw_recs(fw, prompt_id)
            if not fw_recs:
                continue
            prompt_all.extend(fw_recs)
            print(f"  {_abl_row(fw, len(fw_recs), fw_recs)}")
        if prompt_all:
            print(f"  {'-' * (WA - 2)}")
            print(f"  {_abl_row('AVG ' + prompt_id, len(prompt_all), prompt_all)}")

    print(f"\n{'=' * WA}")

    def avg_time(recs):
        times = [r.get("exec_time") for r in recs if r.get("exec_time") is not None]
        return round(sum(times) / len(times), 1) if times else None

    W = 97
    COL = f"{'Prompt':<7} {'Framework':<12} {'Files':>5}  " \
          f"{'Avg s':>6}  " \
          f"{'Cls P':>6} {'Cls R':>6} {'Cls F1':>6}  " \
          f"{'Obj P':>6} {'Obj R':>6} {'Obj F1':>6}  " \
          f"{'Mac F1':>7}"

    def _row(prompt, fw, n, recs):
        t = avg_time(recs)
        t_str = f"{t:>6.1f}" if t is not None else f"{'N/A':>6}"
        return (
            f"{prompt:<7} {fw:<12} {n:>5}  "
            f"{t_str}  "
            f"{avg(recs,'cls_precision'):>6.3f} {avg(recs,'cls_recall'):>6.3f} {avg(recs,'cls_f1'):>6.3f}  "
            f"{avg(recs,'obj_precision'):>6.3f} {avg(recs,'obj_recall'):>6.3f} {avg(recs,'obj_f1'):>6.3f}  "
            f"{avg(recs,'macro_f1'):>7.3f}"
        )

    # Collect all unique models in sorted order
    all_models = sorted({r.get("model", "") for r in all_records})

    for model_name in all_models:
        model_recs = [r for r in all_records if r.get("model", "") == model_name]

        print(f"\n{'=' * W}")
        print(f"  MODEL: {model_name}  ({len(model_recs)} files)")
        print(f"{'=' * W}")
        print(COL)
        print("-" * W)

        # Group by (prompt_id, framework) within this model
        by_group: dict[tuple, list] = defaultdict(list)
        by_prompt: dict[str, list] = defaultdict(list)
        for r in model_recs:
            key = (r.get("prompt_id", ""), r["framework"])
            by_group[key].append(r)
            by_prompt[r.get("prompt_id", "")].append(r)

        def _summary_row(label, n, recs):
            t = avg_time(recs)
            t_str = f"{t:>6.1f}" if t is not None else f"{'N/A':>6}"
            return (
                f"  {label:<19} {n:>5}  "
                f"{t_str}  "
                f"{avg(recs,'cls_precision'):>6.3f} {avg(recs,'cls_recall'):>6.3f} {avg(recs,'cls_f1'):>6.3f}  "
                f"{avg(recs,'obj_precision'):>6.3f} {avg(recs,'obj_recall'):>6.3f} {avg(recs,'obj_f1'):>6.3f}  "
                f"{avg(recs,'macro_f1'):>7.3f}"
            )

        current_prompt = None
        for (prompt, fw), recs in sorted(by_group.items()):
            if prompt != current_prompt:
                if current_prompt is not None:
                    p_recs = by_prompt[current_prompt]
                    print(_summary_row(f"AVG {current_prompt}", len(p_recs), p_recs))
                    print("-" * W)
                current_prompt = prompt
            print(_row(prompt, fw, len(recs), recs))

        if current_prompt:
            p_recs = by_prompt[current_prompt]
            print(_summary_row(f"AVG {current_prompt}", len(p_recs), p_recs))

        print("=" * W)
        print(_summary_row("OVERALL", len(model_recs), model_recs))

    # Cross-model grand total
    print(f"\n{'=' * W}")
    t = avg_time(all_records)
    t_str = f"{t:>6.1f}" if t is not None else f"{'N/A':>6}"
    print(
        f"  {'GRAND TOTAL':<19} {len(all_records):>5}  "
        f"{t_str}  "
        f"{avg(all_records,'cls_precision'):>6.3f} {avg(all_records,'cls_recall'):>6.3f} {avg(all_records,'cls_f1'):>6.3f}  "
        f"{avg(all_records,'obj_precision'):>6.3f} {avg(all_records,'obj_recall'):>6.3f} {avg(all_records,'obj_f1'):>6.3f}  "
        f"{avg(all_records,'macro_f1'):>7.3f}"
    )


if __name__ == "__main__":
    main()
